from selenium.webdriver.support.expected_conditions import staleness_of
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from datetime import date
from time import sleep
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from pathlib import Path
from datetime import datetime
import openpyxl 
from msilib.schema import ServiceControl
from globalConstants import Category
from globalConstants import Product
from globalConstants import Result
from globalConstants import Write
class Test_Localhost:
    vars = {}
    idNumber="0"
    def setup_method(self):
        self.driver = webdriver.Chrome() # webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        now = datetime.now()
        self.folderPath = str(now.strftime("%d-%b-%Y")) #str(date.today())
        self.testTime=str(now.strftime("%H.%M")) #.%S
        Path(self.folderPath).mkdir(exist_ok=True)
        self.driver.get("http://localhost:8080/swagger-ui/index.html#/")
        self.runController(Product.deleteAll,())
        self.runController(Category.deleteAll,())

        
    def teardown_method(self):
        self.vars = {}#sadece test bitiminde sıfırlanır metot bitiminde değil
        self.runController(Product.deleteAll,())
        self.runController(Category.deleteAll,())
        self.driver.quit()



    def getData(excelFileName,sheetName):
        excelFile = openpyxl.load_workbook("data/"+excelFileName)
        selectedSheet = excelFile[sheetName]
        header = {}
        for column in range(1, selectedSheet.max_column + 1):
            header[selectedSheet.cell(row=1, column=column).value] = column
        data = []
        for row in range(2, selectedSheet.max_row + 1):
            row_data = []
            for column in header.values():
                cell_value = selectedSheet.cell(row=row, column=column).value or ""
                row_data.append(cell_value)
            if len(data) > 1:
                data.append(tuple(row_data))
            else:
                data.append(cell_value)
        return data


    def waitForElementVisible(self,locator,timeout=10):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    ######################################################################Make ID##########################################
    def getCategoryId(self):#only for makeCategoryIdNumber
        self.vars["categoryId"]= self.driver.find_element(By.CSS_SELECTOR, ".microlight:nth-child(3) span:nth-child(5)").text
        self.categoryIdNumber= self.vars["categoryId"]
    def getProductId(self):#only for makeProductIdNumber
        self.vars["ProductId"]= self.driver.find_element(By.CSS_SELECTOR, ".microlight:nth-child(3) span:nth-child(10)").text
        self.productIdNumber= self.vars["ProductId"]  

    def makeCategoryIdNumber(self):
        self.runController(Category.deleteAll,())
        self.runController(Category.add,(),Write.body,Write.addCategoryExample)
        self.runController(Category.get,(
            lambda: self.getCategoryId()
        ))
    def makeProductIdNumber(self):
        self.runController(Product.deleteAll,())
        self.makeCategoryIdNumber()
        addProductExaple="{\"description\":\"example\",\"image\":\"example\",\"price\":9.99,\"title\":\"example\","+f"\"categoryId\":{self.categoryIdNumber}"+"}"
        self.runController(Product.add,(),Write.body, addProductExaple)
        self.runController(Product.getAll,(
            lambda: self.getProductId()
        ))

########function list#########
    def result(self,resultNumer):
        self.waitForElementVisible((By.CSS_SELECTOR,".live-responses-table .response > .response-col_status"))
        assert self.driver.find_element(By.CSS_SELECTOR, ".live-responses-table .response > .response-col_status").text == resultNumer
    def errormessage(self,text,position):
        self.waitForElementVisible((By.CSS_SELECTOR, position))
        errormessage = self.driver.find_element(By.CSS_SELECTOR, position)
        assert errormessage.text == f"\"{text}\"" 
#+ save_screenshot





#############################################################runController###########################
    def startController(self,opblockSummaryControl):#only for runController
        self.waitForElementVisible((By.CSS_SELECTOR, opblockSummaryControl))
        self.driver.find_element(By.CSS_SELECTOR, opblockSummaryControl).click()
        self.waitForElementVisible((By.CSS_SELECTOR, ".try-out__btn"))
        self.driver.find_element(By.CSS_SELECTOR, ".try-out__btn").click()

    def stopController(self,opblockSummaryControl):#only for runController
        self.waitForElementVisible((By.CSS_SELECTOR, opblockSummaryControl))
        self.waitForElementVisible((By.CSS_SELECTOR, ".btn-clear"))
        self.waitForElementVisible((By.CSS_SELECTOR, ".btn"))
        self.driver.find_element(By.CSS_SELECTOR, ".btn-clear").click()
        self.driver.find_element(By.CSS_SELECTOR, ".btn").click()
        self.driver.find_element(By.CSS_SELECTOR, opblockSummaryControl).click()

    def callable(self, functions):
        if callable(functions):#eğer fonksiyon tekse fonksiyonu fonksiyon listesine çeviriyoruz
            functions = (functions,)
            return functions
        return functions
    
    def write(self,writingArea,input):
        self.waitForElementVisible((By.CSS_SELECTOR, writingArea))
        bodyParam = self.driver.find_element(By.CSS_SELECTOR, writingArea)
        bodyParam.click()
        bodyParam.clear()
        bodyParam.send_keys(input)

    def runController(self,controller,functions,writingArea=None,input=None):#hiç parametre almıyor
        functions = self.callable(functions)
        self.startController(controller)
        if writingArea and input:#write
            self.write(writingArea,input)
        self.waitForElementVisible((By.CSS_SELECTOR, ".execute"))
        self.driver.find_element(By.CSS_SELECTOR, ".execute").click()
        self.waitForElementVisible((By.CSS_SELECTOR, ".btn-clear"))
        for function in functions:#ekran görüntüsü alma ve test başarısı işlemleri gibi fonksiyonlar
            function()
        self.stopController(controller)  





#############################################################Tests Categories##############################################
    def test_deleteCategories(self):
        self.runController(Category.deleteAll,
            (
            lambda: self.result(Result.ok),
            lambda: self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-deleteCategories.png")
            )
        )

    def test_addCategory(self):
        self.runController(Category.add,
            (
            lambda:self.result(Result.created),
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-addCategory.png")
            ),Write.body,Write.addCategoryExample
        )
        
    @pytest.mark.parametrize("name", getData("category_send_keys.xlsx","CreateCategoryRequest"))
    def test_addCategory_invalid(self,name):
        invalidInput="{"+ f"\"name\":\"{name}\""+"}"
        self.runController(Category.add,
            (
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-addCategory-invalid.png"),
            lambda:self.result(Result.badRequest),
            lambda:self.errormessage("boyut '3' ile '20' arasında olmalı",".language-json > span:nth-child(13)")#hata mesajını excel tablosundan çekmeyi unutma
            ),Write.body,invalidInput
        )

    def test_addCategory_repeat(self):
        self.runController(Category.add,(),Write.body,Write.addCategoryExample)
        self.runController(Category.add,(
            lambda:self.result(Result.badRequest),
            lambda:self.errormessage("Category name already exists",".microlight:nth-child(3) span:nth-child(5)"),
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-addCategory-Repeat.png")
        ),Write.body,Write.addCategoryExample)

    def test_getCategories(self):
        self.runController(Category.get,(
            lambda:self.result(Result.ok),
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getCategories.png")
        ))

    #@pytest.mark.skip()
    def test_updateCategory(self):
        self.makeCategoryIdNumber()
        updateInput="{"+f"\"id\": {self.categoryIdNumber},\"name\": \"example\""+"}"
        self.runController( Category.update,(
            lambda:self.result(Result.ok),
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-updateCategories.png")
        ),Write.body,updateInput)

    def test_deleteCategory(self):
        self.makeCategoryIdNumber()
        self.runController(Category.delete,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-deleteCategory.png"),
            lambda:self.result(Result.ok)
        ),Write.parameters, self.categoryIdNumber)



#################################################################Tests Products########################################
    def test_getAll(self):
        self.runController(Product.getAll,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getAll.png"),
            lambda:self.result(Result.ok)
        ))

    def test_deleteAll(self):
        self.runController(Product.deleteAll, (
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getAll.png"),
            lambda:self.result(Result.ok)
        ))
    
    def test_updateProduct(self):
        self.makeProductIdNumber()
        updateInput="{\"description\":\"example1\",\"id\":"+ self.productIdNumber +",\"image\":\"example1\",\"price\":9.99,\"title\":\"example1\","+f"\"categoryId\":{self.categoryIdNumber}"+"}"
        self.runController(Product.update,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-updateProduct.png"),
            lambda:self.result(Result.ok)
        ),Write.body,updateInput)

    def test_addProduct(self):
        self.makeCategoryIdNumber()
        addProductExaple="{\"description\":\"example\",\"image\":\"example\",\"price\":9.99,\"title\":\"example\","+f"\"categoryId\":{self.categoryIdNumber}"+"}"
        self.runController(Product.add,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-addProduct.png"),
            lambda:self.result(Result.created)
        ),Write.body,addProductExaple)

    def test_getByQueryProductResponse(self):
        self.runController(Product.getByQuery,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getByQueryProductResponse.png"),
            lambda:self.result(Result.ok) 
        ),Write.parameters, Write.queryExample)

    def test_getByCategoryProductResponse(self):
        self.makeCategoryIdNumber()
        self.runController(Product.getByCategory,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getByCategoryProductResponse.png"),
            lambda:self.result(Result.ok)
        ),Write.parameters,self.categoryIdNumber)
    def test_deleteProduct(self):
        self.makeProductIdNumber()
        self.runController(Product.delete,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-deleteProduct.png"),
            lambda:self.result(Result.ok)
        ),Write.parameters,self.productIdNumber)