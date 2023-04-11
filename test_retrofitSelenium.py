from msilib.schema import ServiceControl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import datetime
from datetime import date
import openpyxl 
from selenium.webdriver.support.ui import WebDriverWait as pagewait
from selenium.webdriver.support.expected_conditions import staleness_of
from globalConstants import GlobalConstants as GC

class Test_Localhost:
    vars = {}
    idNumber="0"
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        now = datetime.now()
        self.folderPath = str(now.strftime("%d-%b-%Y")) #str(date.today())
        self.testTime=str(now.strftime("%H.%M")) #.%S
        Path(self.folderPath).mkdir(exist_ok=True)
        self.driver.get("http://localhost:8080/swagger-ui/index.html#/")
    def teardown_method(self):
        self.vars = {}#sadece test bitiminde sıfırlanır metot bitiminde değil
        self.driver.quit()

    def errormessage(self,text,position):
        self.waitForElementVisible((By.CSS_SELECTOR, position))
        errormessage = self.driver.find_element(By.CSS_SELECTOR, position)
        assert errormessage.text == f"\"{text}\"" 

    def getData(excelFileName,sheetName):
        excelFile = openpyxl.load_workbook("data/"+excelFileName)
        selectedSheet = excelFile[sheetName]
        header = {}
        for column in range(1, selectedSheet.max_column + 1):
            header[selectedSheet.cell(row=1, column=column).value] = column
        data = []
        for row in range(2, selectedSheet.max_row + 1):
            row_data = []
            for column in header.values():
                cell_value = selectedSheet.cell(row=row, column=column).value or ""
                row_data.append(cell_value)
            if len(data) > 1:
                data.append(tuple(row_data))
            else:
                data.append(cell_value)
        return data
    
    def result(self,resultNumer):
        self.waitForElementVisible((By.CSS_SELECTOR,".live-responses-table .response > .response-col_status"))
        assert self.driver.find_element(By.CSS_SELECTOR, ".live-responses-table .response > .response-col_status").text == resultNumer

    def waitForElementVisible(self,locator,timeout=10):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    def getCategoryId(self):#self.vars["id"]
        self.vars["categoryId"]= self.driver.find_element(By.CSS_SELECTOR, ".microlight:nth-child(3) span:nth-child(5)").text
        self.categoryIdNumber= self.vars["categoryId"]
        
    def getProductId(self):
        self.vars["ProductId"]= self.driver.find_element(By.CSS_SELECTOR, ".microlight:nth-child(3) span:nth-child(10)").text
        self.productIdNumber= self.vars["ProductId"]    
        

#Controller
    def startController(self,opblockSummaryControl):
        self.waitForElementVisible((By.CSS_SELECTOR, opblockSummaryControl))
        self.driver.find_element(By.CSS_SELECTOR, opblockSummaryControl).click()
        self.waitForElementVisible((By.CSS_SELECTOR, ".try-out__btn"))
        self.driver.find_element(By.CSS_SELECTOR, ".try-out__btn").click()

    def stopController(self,opblockSummaryControl):
        self.waitForElementVisible((By.CSS_SELECTOR, opblockSummaryControl))
        self.waitForElementVisible((By.CSS_SELECTOR, ".btn-clear"))
        self.waitForElementVisible((By.CSS_SELECTOR, ".btn"))
        self.driver.find_element(By.CSS_SELECTOR, ".btn-clear").click()
        self.driver.find_element(By.CSS_SELECTOR, ".btn").click()
        self.driver.find_element(By.CSS_SELECTOR, opblockSummaryControl).click()

    def callable(self, functions):
        if callable(functions):#eğer fonksiyon tekse fonksiyonu fonksiyon listesine çeviriyoruz
            functions = (functions,)
            return functions
        return functions

    def makeCategoryIdNumber(self):
        self.runController(GC.deleteCategories,())
        self.runControllerWithInput(GC.addCategory,GC.body,GC.addCategoryExample,())
        self.runController(GC.getCategories,(
            lambda: self.getCategoryId()
        ))

    def makeProductIdNumber(self):
        self.runController(GC.deleteAll,())
        self.makeCategoryIdNumber()
        addProductExaple="{\"description\":\"example\",\"image\":\"example\",\"price\":9.99,\"title\":\"example\","+f"\"categoryId\":{self.categoryIdNumber}"+"}"
        self.runControllerWithInput(GC.addProduct,GC.body,addProductExaple,())
        self.runController(GC.getAll,(
            lambda: self.getProductId()
        ))
    def write(self,writingArea,input):
        self.waitForElementVisible((By.CSS_SELECTOR, writingArea))
        bodyParam = self.driver.find_element(By.CSS_SELECTOR, writingArea)
        bodyParam.click()
        bodyParam.clear()
        bodyParam.send_keys(input)

#ekle getir gibi fonksiyonlarının yerine body kullanacak ya da kullanmayacak controller işlemini çağırıyoruz.
    def runController(self,controller,functions):#hiç parametre almıyor
        functions = self.callable(functions)
        self.startController(controller)

        self.waitForElementVisible((By.CSS_SELECTOR, ".execute"))
        self.driver.find_element(By.CSS_SELECTOR, ".execute").click()
        self.waitForElementVisible((By.CSS_SELECTOR, ".btn-clear"))
        for function in functions:
            function()

        self.stopController(controller)
    
    def runControllerWithInput(self,controller,writingArea,input,functions):#body ile parametre veriyor
        functions = self.callable(functions)
        self.startController(controller)#GC.controllerAddCategory

        self.write(writingArea,input)

        self.waitForElementVisible((By.CSS_SELECTOR, ".execute"))
        self.driver.find_element(By.CSS_SELECTOR, ".execute").click()
        self.waitForElementVisible((By.CSS_SELECTOR, ".btn-clear"))
        for function in functions:#ekran görüntüsü alma ve test başarısı işlemleri gibi fonksiyonlar
            function()
        self.stopController(controller)

#tests Categories
    def test_deleteCategories(self):
        self.runController(GC.deleteCategories,
            (
            lambda: self.result(GC.ok),
            lambda: self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-deleteCategories.png")
            )
        )

    def test_addCategory(self):
        self.runController(GC.deleteCategories,())
        self.runControllerWithInput(GC.addCategory,GC.body,GC.addCategoryExample,
            (
            lambda:self.result(GC.created),
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-addCategory.png")
            )    
        )
        
    @pytest.mark.parametrize("name", getData("category_send_keys.xlsx","CreateCategoryRequest"))
    def test_addCategory_invalid(self,name):
        self.runController(GC.deleteCategories,())

        invalidInput="{"+ f"\"name\":\"{name}\""+"}"

        self.runControllerWithInput(GC.addCategory,GC.body,invalidInput,
            (
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-addCategory-invalid.png"),
            lambda:self.result(GC.badRequest),
            lambda:self.errormessage("boyut '3' ile '20' arasında olmalı",".language-json > span:nth-child(13)")#hata mesajını excel tablosundan çekmeyi unutma
            )
        )

    def test_addCategory_repeat(self):
        self.runController(GC.deleteCategories,())
        self.runControllerWithInput(GC.addCategory,GC.body,GC.addCategoryExample,())
        self.runControllerWithInput(GC.addCategory,GC.body,GC.addCategoryExample,(
            lambda:self.result(GC.badRequest),
            lambda:self.errormessage("Category name already exists",".microlight:nth-child(3) span:nth-child(5)"),
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-addCategory-Repeat.png")
        ))
        self.runController(GC.deleteCategories,())

    def test_getCategories(self):
        self.runController(GC.getCategories,(
            lambda:self.result(GC.ok),
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getCategories.png")
        ))

    #@pytest.mark.skip()
    def test_updateCategory(self):
        self.makeCategoryIdNumber()
        updateInput="{"+f"\"id\": {self.categoryIdNumber},\"name\": \"example\""+"}"
        self.runControllerWithInput(GC.updateCategory,GC.body,updateInput,(
            lambda:self.result(GC.ok),
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-updateCategories.png")
        ))
        self.runController(GC.deleteCategories,())
    
    def test_deleteCategory(self):
        self.makeCategoryIdNumber()

        self.runControllerWithInput(GC.deleteCategory,GC.parameters,self.categoryIdNumber,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-deleteCategory.png"),
            lambda:self.result(GC.ok)
        ))
        self.runController(GC.deleteCategories,())

#tests Products
    def test_getAll(self):
        self.runController(GC.getAll,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getAll.png"),
            lambda:self.result(GC.ok)
        ))

    def test_deleteAll(self):
        self.runController(GC.deleteAll,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getAll.png"),
            lambda:self.result(GC.ok)
        ))
    
    def test_updateProduct(self):
        self.makeProductIdNumber()
        updateInput="{\"description\":\"example1\",\"id\":"+ self.productIdNumber +",\"image\":\"example1\",\"price\":9.99,\"title\":\"example1\","+f"\"categoryId\":{self.categoryIdNumber}"+"}"
        self.runControllerWithInput(GC.updateProduct,GC.body,updateInput,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-updateProduct.png"),
            lambda:self.result(GC.ok)
        ))
    def test_addProduct(self):
        self.runController(GC.deleteAll,())
        self.makeCategoryIdNumber()
        addProductExaple="{\"description\":\"example\",\"image\":\"example\",\"price\":9.99,\"title\":\"example\","+f"\"categoryId\":{self.categoryIdNumber}"+"}"
        self.runControllerWithInput(GC.addProduct,GC.body,addProductExaple,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-addProduct.png"),
            lambda:self.result(GC.created)
        ))

    def test_getByQueryProductResponse(self):
        self.runController(GC.deleteAll,())

        self.runControllerWithInput(GC.getByQueryProductResponse,GC.parameters,"example",(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getByQueryProductResponse.png"),
            lambda:self.result(GC.ok) 
        ))

    def test_getByCategoryProductResponse(self):
        self.runController(GC.deleteAll,())
        self.makeCategoryIdNumber()
        self.runControllerWithInput(GC.getByCategoryProductResponse,GC.parameters,self.categoryIdNumber,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-getByCategoryProductResponse.png"),
            lambda:self.result(GC.ok) 
        ))
    def test_deleteProduct(self):
        self.runController(GC.deleteAll,())
        self.makeProductIdNumber()
        self.runControllerWithInput(GC.deleteProduct,GC.parameters,self.productIdNumber,(
            lambda:self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-retrofitSelenium-deleteProduct.png"),
            lambda:self.result(GC.ok) 
        ))
